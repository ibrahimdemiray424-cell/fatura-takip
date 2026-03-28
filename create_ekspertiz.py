from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

wb = Workbook()
ws = wb.active
ws.title = "Ön Hasar Ekspertiz"

# --- Stil tanımları ---
def thin_border(top=True, bottom=True, left=True, right=True):
    s = Side(style="thin", color="000000")
    n = Side(style=None)
    return Border(
        top=s if top else n,
        bottom=s if bottom else n,
        left=s if left else n,
        right=s if right else n,
    )

def thick_border():
    s = Side(style="medium", color="000000")
    return Border(top=s, bottom=s, left=s, right=s)

COLOR_HEADER_BG   = "1F3864"  # koyu lacivert
COLOR_HEADER_FG   = "FFFFFF"
COLOR_SECTION_BG  = "BDD7EE"  # açık mavi
COLOR_TABLE_HDR   = "2E75B6"  # mavi
COLOR_TABLE_HDR_FG= "FFFFFF"
COLOR_ROW_ALT     = "DEEAF1"  # zebra
COLOR_WHITE       = "FFFFFF"
COLOR_LABEL_BG    = "F2F2F2"

def style_cell(cell, bold=False, size=10, fg=None, bg=None,
               h_align="left", v_align="center", wrap=False, border=None):
    cell.font = Font(name="Calibri", bold=bold, size=size,
                     color=fg if fg else "000000")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align,
                                wrap_text=wrap)
    if border:
        cell.border = border

def merge_style(ws, cell_range, value="", bold=False, size=10,
                fg=None, bg=None, h_align="left", v_align="center",
                wrap=False, border=None):
    ws.merge_cells(cell_range)
    first = ws[cell_range.split(":")[0]]
    first.value = value
    style_cell(first, bold=bold, size=size, fg=fg, bg=bg,
               h_align=h_align, v_align=v_align, wrap=wrap, border=border)
    # kenarlık tüm birleşik hücrelere
    if border:
        for row in ws[cell_range]:
            for c in row:
                c.border = border

# --- Sütun genişlikleri ---
# A  B  C  D  E  F  G  H  I
col_widths = [2, 22, 14, 14, 14, 12, 12, 12, 22]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Satır yükseklikleri (sonra tek tek ayarlanacak) ---
row = 1

# ===================== BAŞLIK =====================
ws.row_dimensions[row].height = 8
row += 1  # row 2

ws.row_dimensions[row].height = 36
merge_style(ws, f"B{row}:I{row}",
            value="ARAÇ ÖN HASAR EKSPERTİZ FORMU",
            bold=True, size=16, fg=COLOR_HEADER_FG, bg=COLOR_HEADER_BG,
            h_align="center", v_align="center",
            border=thin_border())
row += 1  # row 3

ws.row_dimensions[row].height = 18
# Rapor No / Tarih / Ekspertiz Yeri
for col, label, val_col in [("B", "Rapor No :", "C"), ("E", "Tarih :", "F"), ("G", "Ekspertiz Yeri :", "I")]:
    c = ws[f"{col}{row}"]
    c.value = label
    style_cell(c, bold=True, size=9, bg=COLOR_LABEL_BG, border=thin_border())
    vc = ws[f"{val_col}{row}"]
    style_cell(vc, size=9, border=thin_border())

ws.merge_cells(f"C{row}:D{row}")
ws.merge_cells(f"F{row}:F{row}")
ws.merge_cells(f"I{row}:I{row}")
row += 1  # row 4
ws.row_dimensions[row].height = 6
row += 1  # row 5

# ===================== ARAÇ BİLGİLERİ =====================
ws.row_dimensions[row].height = 18
merge_style(ws, f"B{row}:I{row}",
            value="  ARAÇ BİLGİLERİ",
            bold=True, size=10, fg="000000", bg=COLOR_SECTION_BG,
            h_align="left", v_align="center",
            border=thin_border())
row += 1

arac_fields = [
    [("Plaka", "C"), ("Marka / Model", "F")],
    [("Yıl", "C"), ("Renk", "F")],
    [("Şase No", "C"), ("Motor No", "F")],
    [("Kilometre", "C"), ("Yakıt Türü", "F")],
]

for pair in arac_fields:
    ws.row_dimensions[row].height = 18
    for label, val_end in pair:
        if val_end == "C":
            lbl_col, val_start, val_end2 = "B", "C", "D"
        else:
            lbl_col, val_start, val_end2 = "E", "F", "I"
        lc = ws[f"{lbl_col}{row}"]
        lc.value = label
        style_cell(lc, bold=True, size=9, bg=COLOR_LABEL_BG, border=thin_border())
        ws.merge_cells(f"{val_start}{row}:{val_end2}{row}")
        vc = ws[f"{val_start}{row}"]
        style_cell(vc, size=9, border=thin_border())
    row += 1

ws.row_dimensions[row].height = 6
row += 1

# ===================== SİGORTA BİLGİLERİ =====================
ws.row_dimensions[row].height = 18
merge_style(ws, f"B{row}:I{row}",
            value="  SİGORTA / POLİÇE BİLGİLERİ",
            bold=True, size=10, fg="000000", bg=COLOR_SECTION_BG,
            h_align="left", v_align="center",
            border=thin_border())
row += 1

sig_fields = [
    [("Sigorta Şirketi", "C"), ("Poliçe No", "F")],
    [("Hasar Tarihi", "C"), ("İhbar Tarihi", "F")],
    [("Dosya No", "C"), ("", "F")],
]

for pair in sig_fields:
    ws.row_dimensions[row].height = 18
    for label, val_end in pair:
        if val_end == "C":
            lbl_col, val_start, val_end2 = "B", "C", "D"
        else:
            lbl_col, val_start, val_end2 = "E", "F", "I"
        lc = ws[f"{lbl_col}{row}"]
        lc.value = label
        style_cell(lc, bold=True, size=9, bg=COLOR_LABEL_BG, border=thin_border())
        ws.merge_cells(f"{val_start}{row}:{val_end2}{row}")
        vc = ws[f"{val_start}{row}"]
        style_cell(vc, size=9, border=thin_border())
    row += 1

ws.row_dimensions[row].height = 6
row += 1

# ===================== HASAR TESPİT TABLOSU =====================
ws.row_dimensions[row].height = 18
merge_style(ws, f"B{row}:I{row}",
            value="  HASAR TESPİT TABLOSU (ÖN BÖLGE)",
            bold=True, size=10, fg="000000", bg=COLOR_SECTION_BG,
            h_align="left", v_align="center",
            border=thin_border())
row += 1

# Tablo başlıkları
tbl_headers = ["PARÇA", "HASAR YOK", "BOYALI", "DEĞİŞMİŞ", "HASARLI", "KOPUK/\nKIRIK", "NOTLAR"]
tbl_cols    = ["B", "C", "D", "E", "F", "G", "H"]  # H->I birleştirilecek
ws.row_dimensions[row].height = 28
for i, (col, hdr) in enumerate(zip(tbl_cols, tbl_headers)):
    c = ws[f"{col}{row}"]
    c.value = hdr
    style_cell(c, bold=True, size=9, fg=COLOR_TABLE_HDR_FG, bg=COLOR_TABLE_HDR,
               h_align="center", v_align="center", wrap=True, border=thin_border())
# Notlar sütunu B ile I arasında H:I
ws.merge_cells(f"H{row}:I{row}")
row += 1

parcalar = [
    "Ön Tampon",
    "Kaput",
    "Sağ Ön Çamurluk",
    "Sol Ön Çamurluk",
    "Ön Izgara",
    "Sağ Far",
    "Sol Far",
    "Sağ Ön Kapı",
    "Sol Ön Kapı",
    "Ön Cam",
    "Motor (Görsel)",
    "Radyatör (Görsel)",
    "Airbag",
    "Diğer",
]

for idx, parca in enumerate(parcalar):
    ws.row_dimensions[row].height = 16
    bg = COLOR_ROW_ALT if idx % 2 == 0 else COLOR_WHITE
    # Parça adı
    pc = ws[f"B{row}"]
    pc.value = parca
    style_cell(pc, size=9, bg=bg, border=thin_border())
    # Checkbox sütunlar (C-G) — boş, ortalı
    for col in ["C", "D", "E", "F", "G"]:
        c = ws[f"{col}{row}"]
        style_cell(c, size=9, bg=bg, h_align="center", border=thin_border())
    # Notlar (H:I)
    ws.merge_cells(f"H{row}:I{row}")
    nc = ws[f"H{row}"]
    style_cell(nc, size=9, bg=bg, border=thin_border())
    row += 1

ws.row_dimensions[row].height = 6
row += 1

# ===================== EKSPERTİZ NOTLARI =====================
ws.row_dimensions[row].height = 18
merge_style(ws, f"B{row}:I{row}",
            value="  EKSPERTİZ NOTLARI / GENEL GÖZLEMLER",
            bold=True, size=10, fg="000000", bg=COLOR_SECTION_BG,
            h_align="left", v_align="center",
            border=thin_border())
row += 1

ws.row_dimensions[row].height = 60
merge_style(ws, f"B{row}:I{row}",
            value="",
            bold=False, size=9, bg=COLOR_WHITE,
            h_align="left", v_align="top", wrap=True,
            border=thin_border())
row += 1

ws.row_dimensions[row].height = 6
row += 1

# ===================== İMZA BÖLÜMÜ =====================
ws.row_dimensions[row].height = 18
merge_style(ws, f"B{row}:I{row}",
            value="  İMZA",
            bold=True, size=10, fg="000000", bg=COLOR_SECTION_BG,
            h_align="left", v_align="center",
            border=thin_border())
row += 1

ws.row_dimensions[row].height = 50
# Araç Sahibi
ws.merge_cells(f"B{row}:E{row}")
ac = ws[f"B{row}"]
ac.value = "Araç Sahibi\nAd Soyad :\nİmza         :"
style_cell(ac, size=9, bg=COLOR_LABEL_BG, h_align="left", v_align="top",
           wrap=True, border=thin_border())
# Ekspertiz Uzmanı
ws.merge_cells(f"F{row}:I{row}")
ec = ws[f"F{row}"]
ec.value = "Ekspertiz Uzmanı\nAd Soyad :\nÜnvan        :\nİmza           :"
style_cell(ec, size=9, bg=COLOR_LABEL_BG, h_align="left", v_align="top",
           wrap=True, border=thin_border())
row += 1

ws.row_dimensions[row].height = 8

# ===================== YAZDIRILACAK ALAN =====================
last_data_row = row
ws.print_area = f"A1:{get_column_letter(9)}{last_data_row}"

# ===================== SAYFA AYARLARI =====================
ws.page_setup.paperSize   = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 1

ws.page_margins = PageMargins(
    left=0.4, right=0.4, top=0.5, bottom=0.5,
    header=0.2, footer=0.2
)

ws.sheet_properties.pageSetUpPr.fitToPage = True

# ===================== KAYDET =====================
out_path = r"C:\Users\idemi\Desktop\verdent-project\on_hasar_ekspertiz.xlsx"
wb.save(out_path)
print(f"Dosya kaydedildi: {out_path}")
